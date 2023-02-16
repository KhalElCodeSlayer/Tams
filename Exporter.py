import itertools

import openpyxl as op
import os
from Imp import Imp
from aircraft import Operation
from aircraft import Aircraft
from aircraft import Component
from aircraft import Componento


class Exporter:

    def __init__(self,aircraft: Aircraft):
        self.create(aircraft)
        self.expAir(aircraft)
        self.expCompo(aircraft)
        self.expOp(aircraft)
        self._wb.save("Data.xlsx")
        self.expm_rec(aircraft, "2021")


    def create(self, aircraft: Aircraft):
        info = ["Aircraft Information", "Operations", "Component Maintenance Record", "Component Maintenance Rec"]
        self._wb = op.Workbook()
        for i in info:
            self._wb.create_sheet(title=i)

        self._wb.remove(self._wb['Sheet'])
        self._wb.save("Data.xlsx")

    def expAir(self, aircraft: Aircraft):
        sheet = self._wb["Aircraft Information"]
        for x in range(0, len(list(aircraft.get_dict().keys()))):
            sheet.cell(row=x + 1, column=1).value = list(aircraft.get_dict().keys())[x]
            sheet.cell(row=x + 1, column=2).value = aircraft.get_dict()[list(aircraft.get_dict().keys())[x]]

    def expCompo(self, aircraft: Aircraft):
        sheet = self._wb["Component Maintenance Rec"]
        compo = aircraft.get_info("Componentso")
        keys = list(compo.keys())
        for x in range(0, len(keys)):
            compdict = compo[list(compo.keys())[x]].toString()
            for i in range(0, len(list(compdict.keys()))):
                sheet.cell(row=x + 4, column=i + 1).value = compdict[list(compdict.keys())[i]]

    def expOp(self, aircraft: Aircraft):
        sheet = self._wb["Operations"]
        operations = aircraft.get_info("Operations")
        keys = list(operations.keys())
        for x in range(0, len(keys)):
            opdict = operations[list(operations.keys())[x]].toString()
            for i in range(0, len(opdict.keys())):
                sheet.cell(row=x + 3, column=i + 1).value = opdict[list(opdict.keys())[i]]


    def expm_rec(self,aircraft: Aircraft,year):
        #change to be compatible with multiple years
        crec = aircraft.crec
        orec = aircraft.orec
        months = ["JAN", "FEB", "MAR", "APRIL", "MAY", "JUNE", "JULY", "AUG", "SEPT", "OCT", "NOV", "DEC"]
        print(crec)
        for i in months:
            for component,operation in itertools.zip_longest(crec[i],orec[i]):
                if component != None:
                    cfile = open(f"Airplane/Components/{year}/{i}/{component}.txt", "w")
                    tempc = crec[i][component]
                    cstring = ''
                    if tempc != {}:
                        for k in tempc:
                            tempcc = tempc[k]
                            tempcc = list(tempcc.values())
                            cstring += f"{k}:{tempcc[0]}:{tempcc[1]}:{tempcc[2]}\n"
                    cfile.write(cstring)
                    cfile.close()
                if operation != None:
                    ofile = open(f"Airplane/Operations/{year}/{i}/{operation}.txt", "w")
                    tempo = orec[i][operation]
                    ostring = ''
                    if tempo != {}:
                        for k in tempo:
                            tempoo = tempo[k]
                            tempoo = list(tempoo.values())
                            ostring += f"{k}:{tempoo[0]}:{tempoo[1]}:{tempoo[2]}\n"
                    ofile.write(cstring)
                    ofile.close()




    def stubby(self):
        pass


if __name__ == "__main__":
    Pt=os.path.join(os.getcwd(),"Data.xlsx")
    a = Imp(Pt)
    ex = Exporter("Data.xlsx", a.get_Aircraft())
