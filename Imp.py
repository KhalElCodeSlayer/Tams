import openpyxl as op
import os
from datetime import date
from FileMng import FileMng
from aircraft import Operation
from aircraft import Aircraft
from aircraft import Component
from aircraft import Componento


class Imp:
    Aircraft_info = {}
    Operations = {}
    Component = {}
    Componento = {}

    def __init__(self, path):
        Workbook = op.load_workbook(path, data_only=True)
        sheet = Workbook["Aircraft Information"]
        validsheets = ["Aircraft Information", "Operations", "Component Maintenance Record",
                       "Component Maintenance Rec"]
        for sheetname in validsheets:
            if sheetname in Workbook.sheetnames:
                if sheetname == validsheets[0]:
                    self.iAircraft(Workbook[sheetname])
                elif sheetname == validsheets[1]:
                    self.iOperation(Workbook[sheetname])
                elif sheetname == validsheets[2]:
                    self.iComponent(Workbook[sheetname])
                elif sheetname == validsheets[3]:
                    self.iComponento(Workbook[sheetname])
        self.aircraft = Aircraft.Aircraft(self.Aircraft_info, self.Component, self.Componento, self.Operations)
        if not os.path.exists("Airplane"):
            FileMng(self.aircraft,date.year)
            #set this up to use multiple years, but for now i cant bother
        self.aircraft.init_mrec(self.i_cyear(2021),self.i_oyear(2021))

    def iAircraft(self, sheet):
        attr = ["Aircraft Make", "Model", "Serial #", "Registration #", "Date made", "Hobbs adjust", "Tach", "Hobbs",
                "Next Insp", "Propeller T.T.", "Prop.  SMOH", "NEXT INSP DUE", "Engine T.T.", "Eng. SMOH",
                "Time Remaining To Next Insp.","M_Cost"]
        for row_cells in sheet.iter_rows():
            for cell in range(len(row_cells)):
                if row_cells[cell].value in attr:
                    self.Aircraft_info[row_cells[cell].value] = row_cells[cell + 1].value
        # done

    def iOperation(self, sheet):
        op = {"Inspection": None, "Frequency": None, "Hours": None, "Months": None, "HCW": None, "DCW": None,
              "C_Tach": None, "NDH": None, "NDD": None, "Note": None, "T_rem": None}
        c = 1
        for row_cells in sheet.iter_rows():
            new = op.copy()
            if 'Operation' in str(row_cells[0].value):
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    try:
                        new[list(new.keys())[cell]] = row_cells[cell].value
                    except:
                        pass
                self.Operations[id] = Operation.Operation(new)
                c += 1
        # done

    def iComponent(self, sheet):
        comp = {"Item": None, "P_No": None, "S_No": None, "Date": None, "RY": None, "RM": None, "RH": None, "RO": None,
                "Tot_I": None, "Past_T": None, "Comp_cycle": None, "Date_CW": None, "T_Remove": None,
                "Sched_Remove": None, "T_RemH": None, "T_remD": None}
        c = 0
        for row_cells in sheet.iter_rows():
            new = comp.copy()
            if str(row_cells[0].value) != "Item" and str(row_cells[0].value) != "None" and c > 2:
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    try:
                        new[list(new.keys())[cell]] = row_cells[cell].value
                    except:
                        pass
                self.Component[id] = Component.Component(new)
            c += 1
        # done?

    def iComponento(self, sheet):
        compo = {"Item": None, "P_No": None, "S_No": None, "Date": None, "RY": None, "RH": None, "Tot_I": None,
                 "Past_T": None, "Curr_time": None, "AF_Hours": None, "L_Rem": None, "Date_rep": None,
                 "Next_inspH": None, "Next_insp": None, "Type_mjr": None, "Time_mjr": None,"Remarks": None,"M_Cost":None}
        c = 0
        for row_cells in sheet.iter_rows():
            new = compo.copy()

            if str(row_cells[0].value) != "Item" and str(row_cells[0].value) != "None" and c > 2:
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    try:
                        new[list(new.keys())[cell]] = row_cells[cell].value
                    except:
                        pass
                self.Componento[id] = Componento.Componento(new)
            c += 1

    def i_cyear(self,year):
        ylist = os.listdir(f"Airplane/Components")
        ydict = dict(zip(os.listdir(f"Airplane/Components/{year}"), [None for i in range(len(os.listdir(f"Airplane/Components/{year}")))]))
        if str(year) in ylist:
            for i in os.listdir(f"Airplane/Components/{year}"):
                ydict[i] = self.i_cmonth(i,year)
        return ydict

    def i_cmonth(self,month,year):
        mlist = os.listdir(f"Airplane/Components/{year}")
        mdict = dict(zip([p[:-4] for p in  os.listdir(f"Airplane/Components/{year}/{month}")],[None for i in range(len(os.listdir(f"Airplane/Components/{year}/{month}")))]))
        if month in mlist:
            for i in os.listdir(f"Airplane/Components/{year}/{month}"):
                mdict[i[:-4]] = self.i_ctxt(month,year,i[:-4])
        return mdict

    def i_ctxt(self,month,year,component):
        cfile = open(f"Airplane/Components/{year}/{month}/{component}.txt","r")
        clines = cfile.readlines()
        datalist = ["ID","Cost","Comments"]
        clines = [i.rstrip("\n").split(':') for i in clines]
        cdict = dict(zip([i[0] for i in clines], [None for k in range(100)]))
        rdict = {}
        for j in clines:
            cdict[j[0]] = dict(zip(datalist,j[1:]))
        cfile.close()
        return cdict


    def i_oyear(self,year):
        ylist = os.listdir(f"Airplane/Operations")
        ydict = dict(zip(os.listdir(f"Airplane/Operations/{year}"), [None for i in range(len(os.listdir(f"Airplane/Operations/{year}")))]))
        if str(year) in ylist:
            for i in os.listdir(f"Airplane/Components/{year}"):
                ydict[i] = self.i_omonth(i,year)
        return ydict

    def i_omonth(self,month,year):
        mlist = os.listdir(f"Airplane/Operations/{year}")
        mdict = dict(zip([p[:-4] for p in  os.listdir(f"Airplane/Operations/{year}/{month}")],[None for i in range(len(os.listdir(f"Airplane/Operations/{year}/{month}")))]))
        if month in mlist:
            for i in os.listdir(f"Airplane/Operations/{year}/{month}"):
                mdict[i[:-4]] = self.i_otxt(month,year,i[:-4])
        return mdict

    def i_otxt(self,month,year,component):
        ofile = open(f"Airplane/Operations/{year}/{month}/{component}.txt","r")
        olines = ofile.readlines()
        datalist = ["ID","Cost","Comments"]
        olines = [i.rstrip("\n").split(':') for i in olines]
        odict = dict(zip([i[0] for i in olines], [None for k in range(100)]))
        for j in olines:
            odict[j[0]] = dict(zip(datalist,j[1:]))
        ofile.close()
        return odict




    def get_Aircraft(self):
        return self.aircraft





